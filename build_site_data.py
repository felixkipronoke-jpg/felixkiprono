import csv
import json
import re
from collections import OrderedDict
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parent
PORTFOLIO_CSV = ROOT / "full_portfolio.csv"
SOCIAL_CSV = ROOT / "social_videos.csv"
DATA_JS = ROOT / "site-data.js"
TEMPLATE_XLSX = ROOT / "website_content_template.xlsx"
CONTENT_XLSX = ROOT / "website_content.xlsx"

PLATFORM_PRIORITY = {"youtube": 0, "tiktok": 1, "instagram": 2}

TOPIC_RULES = OrderedDict(
    [
        ("health", ["health", "healthcare", "hiv", "malaria", "maternal", "nutrition", "usaid"]),
        ("gender", ["femicide", "women", "gender", "gbv", "justice"]),
        ("economics", ["debt", "finance", "bill", "petrol", "electricity", "inflation", "tax", "bank", "price"]),
        ("environment", ["forest", "climate", "environment", "geothermal", "landscape", "volcano", "nyiragongo"]),
        ("governance", ["corrupt", "parliament", "public", "accountability", "government", "civic"]),
        ("migration", ["migrant", "migration", "tourism", "visiting"]),
        ("education", ["school", "students", "education"]),
        ("media", ["storytelling", "journalism", "media", "showreel", "video"]),
    ]
)

PROFILE_ROWS = [
    {"field": "hero_label", "value": "Data Journalist"},
    {"field": "hero_title", "value": "I turn public data into stories people can use."},
    {
        "field": "hero_subtitle",
        "value": "I build evidence-led stories, databases, explainers, and videos across public finance, health, climate, gender, economics, and civic accountability.",
    },
    {"field": "about_title", "value": "Curiosity, public evidence, and simple explanation guide the work."},
    {
        "field": "about_paragraph_1",
        "value": "I love looking at data, spotting trends, and making meaningful connections. My work turns those connections into useful public knowledge, using clear language and visual storytelling so more people can understand the world around them.",
    },
    {
        "field": "about_paragraph_2",
        "value": "I started in finance and accounting, where systems, controls, and numbers shaped how I think. Moving into journalism was also a decision to pursue my passion: finding what is hidden in plain sight and explaining it in ways that can inspire innovation and better decisions.",
    },
    {
        "field": "about_paragraph_3",
        "value": "I believe readily available tools can produce remarkable public-interest work when used with care. My journey into journalism and the skills I have built reflect that belief: collect what is public, make it accessible, and help people see what the evidence is saying.",
    },
    {"field": "contact_intro", "value": "For data stories, investigations, training, partnerships, and editorial projects."},
    {"field": "email", "value": "felixkiprono.ke@gmail.com"},
]

EXPERIENCE_ROWS = [
    {
        "period": "Apr 2023 - Present",
        "title": "Data journalist and editor",
        "organization": "Odipo Dev / Africa Data Hub",
        "summary": "Leads data storytelling work and content editing across a consortium of four organizations, producing public-interest stories and helping build an Africa-wide data repository. Key work includes the Silencing Women femicide database and award-winning reporting on the effect of US funding withdrawals on Kenya's health system.",
        "bullets": "Sources, extracts, cleans, analyzes, and quality-checks public data from the World Bank, OECD, UN, IMF, national statistics bureaus, and other repositories.|Turns complex datasets into articles, explainers, reports, videos, and visual stories for non-technical audiences.|Maintains editorial and data quality standards across Africa Data Hub and Odipo Dev projects.",
    },
    {
        "period": "Jun 2021 - Oct 2022",
        "title": "Multimedia journalist",
        "organization": "Debunk Media",
        "summary": "Reported and produced financial and economic stories on public debt, inflation, and civic issues, packaging them into accessible multimedia explainers and evergreen social content.",
        "bullets": "Researched public-interest topics, gathered data, mined insights, and wrote video scripts.|Produced articles, videos, graphics, animations, and social media content aligned with Debunk Media editorial standards.",
    },
    {
        "period": "Jul 2019 - Sep 2021",
        "title": "Cash and bank accountant",
        "organization": "MultiChoice Africa",
        "summary": "Prepared payment statistics, processed transaction matching, completed monthly bank reconciliations, and strengthened financial controls to reduce revenue leakage.",
        "bullets": "",
    },
    {
        "period": "Jul 2013 - Jun 2019",
        "title": "Treasury accountant",
        "organization": "MultiChoice Africa",
        "summary": "Performed bank reconciliations in line with accounting standards and prepared monthly cash analysis reports for management.",
        "bullets": "",
    },
]

EDUCATION_ROWS = [
    {
        "period": "Sep 2009 - Apr 2013",
        "title": "Bachelor's degree in Supply Chain Management",
        "institution": "Jomo Kenyatta University of Agriculture and Technology",
        "details": "",
    },
    {
        "period": "Jul 2008 - Jun 2012",
        "title": "Certified Public Accountant (CPA-K)",
        "institution": "Kenya College of Accountancy and self-study",
        "details": "",
    },
]

SKILL_ROWS = [
    {"skill": "CPA-K", "details": ""},
    {"skill": "Data analysis", "details": ""},
    {"skill": "Data visualization", "details": ""},
    {"skill": "Content writing", "details": ""},
    {"skill": "Editing", "details": ""},
    {"skill": "Publishing", "details": ""},
    {"skill": "Web scraping", "details": ""},
    {"skill": "Document extraction", "details": ""},
    {"skill": "Remote teamwork", "details": ""},
    {"skill": "Time management", "details": ""},
    {"skill": "MS Excel", "details": ""},
    {"skill": "Google Sheets", "details": ""},
    {"skill": "OpenRefine", "details": ""},
    {"skill": "Google Data Studio", "details": ""},
    {"skill": "Datawrapper", "details": ""},
    {"skill": "Flourish", "details": ""},
    {"skill": "SQL Server", "details": ""},
    {"skill": "Python", "details": ""},
    {"skill": "Google Pinpoint", "details": ""},
    {"skill": "Claude", "details": ""},
    {"skill": "LLMs", "details": ""},
    {"skill": "MS Office", "details": ""},
    {"skill": "Google Suite", "details": ""},
]

PARTNER_ROWS = [
    {"name": "Odipo Dev", "url": "https://www.odipodev.com/", "logo_path": "assets/partners/odipo-dev.png", "category": "Organization"},
    {"name": "AIJC", "url": "https://aijc.africa/", "logo_path": "assets/partners/aijc.png", "category": "Conference"},
    {"name": "GIJN", "url": "https://gijn.org/", "logo_path": "assets/partners/gijn.png", "category": "Network"},
    {"name": "Debunk Media", "url": "https://debunk.media/", "logo_path": "assets/partners/debunk-media.png", "category": "Media"},
    {"name": "Baraza Media Lab", "url": "https://barazalab.com/", "logo_path": "assets/partners/baraza-media-lab.png", "category": "Media lab"},
    {"name": "Africa Uncensored", "url": "https://africauncensored.online/", "logo_path": "assets/partners/africa-uncensored.png", "category": "Media"},
    {"name": "Nation Media Group", "url": "https://nation.africa/", "logo_path": "assets/partners/nation-media-group.png", "category": "Media"},
    {"name": "DW Akademie", "url": "https://www.dw.com/en/dw-akademie/s-12130", "logo_path": "assets/partners/dw-akademie.png", "category": "Training"},
]

SOCIAL_ROWS = [
    {"platform": "Instagram", "url": "https://www.instagram.com/kipronoexplores?igsh=MTZrNmllOXYzd2s2YQ=="},
    {"platform": "TikTok", "url": "https://www.tiktok.com/@felixkiprono.ke?_r=1&_t=ZS-960QqsIFACS"},
    {"platform": "X", "url": "https://x.com/felixkiprono_ke"},
    {"platform": "YouTube", "url": "https://www.youtube.com/@kipronoexplores1981"},
]


def read_csv(path):
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return [clean_row(row) for row in csv.DictReader(handle)]


def clean_row(row):
    return {str(k or "").strip(): str(v or "").strip() for k, v in row.items()}


def rows_from_sheet(workbook, sheet_name, defaults):
    if sheet_name not in workbook.sheetnames:
        return defaults
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return defaults
    headers = [str(value or "").strip() for value in rows[0]]
    data = []
    for row in rows[1:]:
        entry = {headers[index]: str(value or "").strip() for index, value in enumerate(row) if index < len(headers)}
        if any(entry.values()):
            data.append(entry)
    return data or defaults


def load_site_content():
    if CONTENT_XLSX.exists():
        workbook = load_workbook(CONTENT_XLSX, data_only=True)
        return {
            "profile": {row.get("field", ""): row.get("value", "") for row in rows_from_sheet(workbook, "profile", PROFILE_ROWS)},
            "experience": rows_from_sheet(workbook, "experience", EXPERIENCE_ROWS),
            "education": rows_from_sheet(workbook, "education", EDUCATION_ROWS),
            "skills": rows_from_sheet(workbook, "skills", SKILL_ROWS),
            "partners": rows_from_sheet(workbook, "partners", PARTNER_ROWS),
            "socials": rows_from_sheet(workbook, "socials", SOCIAL_ROWS),
        }

    return {
        "profile": {row["field"]: row["value"] for row in PROFILE_ROWS},
        "experience": EXPERIENCE_ROWS,
        "education": EDUCATION_ROWS,
        "skills": SKILL_ROWS,
        "partners": PARTNER_ROWS,
        "socials": SOCIAL_ROWS,
    }


def normalize_platform(value):
    value = (value or "").strip()
    lower = value.lower()
    if "youtube" in lower:
        return "YouTube"
    if "tiktok" in lower:
        return "TikTok"
    if "instagram" in lower:
        return "Instagram"
    return value


def source_key(row):
    source_id = row.get("source_id", "").strip().lower()
    title = row.get("title", "").strip().lower()
    url = row.get("url") or row.get("video_url") or ""
    parsed = urlparse(url)
    path = parsed.path.rstrip("/").lower()

    if source_id:
        return source_id
    if path:
        return path
    return re.sub(r"[^a-z0-9]+", "-", title).strip("-")


def title_key(row):
    title = row.get("title", "").strip().lower()
    normalized = re.sub(r"[^a-z0-9]+", " ", title).strip()
    if len(normalized) < 18:
        return ""
    return normalized


def platform_rank(row):
    return PLATFORM_PRIORITY.get(normalize_platform(row.get("platform", "")).lower(), 9)


def infer_topics(row):
    provided = row.get("topic", "")
    if provided:
        topics = [re.sub(r"[^a-z0-9]+", "-", part.strip().lower()).strip("-") for part in re.split(r"[,;/]", provided)]
        return [topic for topic in topics if topic] or ["data"]
    haystack = " ".join([row.get("title", ""), row.get("summary", ""), row.get("behind_the_story", "")]).lower()
    topics = [topic for topic, words in TOPIC_RULES.items() if any(word in haystack for word in words)]
    return topics or ["data"]


def infer_year(date):
    match = re.search(r"(20\d{2})", date or "")
    return match.group(1) if match else "Undated"


def fallback_thumbnail(row):
    platform = normalize_platform(row.get("platform", ""))
    url = row.get("url") or row.get("video_url") or ""
    source_id = row.get("source_id", "")
    if platform == "YouTube" and source_id:
        return f"https://i.ytimg.com/vi/{source_id}/hqdefault.jpg"
    if "africadatahub.org" in url:
        return "https://cdn.prod.website-files.com/609a29dfdde412503cff3e96/65ae27e01d266d9792150286_65ab49db5648e593287bdee3_femicide_collage.jpg"
    if "bbc.com" in url:
        return "https://ichef.bbci.co.uk/news/1024/branded_news/3EF3/production/_132551161_902fb94d13d63145f38f719ac849bda55684f73f0_0_6720_44801000x667.jpg"
    if "willowhealthmedia.org" in url:
        return "https://willowhealthmedia.org/storage/2025/05/doctor-performing-medical-checkup-patient-scaled.jpg"
    if "debunk.media" in url:
        return "https://debunk.media/wp-content/uploads/Artboard-2-100-6.jpg"
    return "https://images.unsplash.com/photo-1551288049-bebda4e38f71?auto=format&fit=crop&w=1200&q=80"


def to_item(row):
    content_type = (row.get("content_type") or "video").strip().lower()
    platform = normalize_platform(row.get("platform", ""))
    url = row.get("url") or row.get("video_url") or ""
    publisher = row.get("publisher_client") or row.get("account") or platform
    role = row.get("role", "")

    if content_type == "video" and not role:
        role = "Creator / producer"

    item = {
        "type": content_type,
        "platform": platform,
        "account": row.get("account", ""),
        "title": row.get("title", "Untitled"),
        "date": row.get("date", ""),
        "year": infer_year(row.get("date", "")),
        "summary": row.get("summary", ""),
        "url": url,
        "thumbnail": row.get("thumbnail_url") or fallback_thumbnail(row),
        "sourceId": row.get("source_id", ""),
        "role": role,
        "publisher": publisher,
        "priority": row.get("priority_level", "low priority").lower(),
        "story": row.get("behind_the_story", ""),
    }
    item["topics"] = infer_topics(item)
    return item


def load_items():
    rows = read_csv(PORTFOLIO_CSV)
    for social in read_csv(SOCIAL_CSV):
        rows.append(
            {
                "content_type": "video",
                "platform": social.get("platform", ""),
                "account": social.get("account", ""),
                "title": social.get("title", ""),
                "date": social.get("date", ""),
                "summary": social.get("summary", ""),
                "url": social.get("video_url", ""),
                "thumbnail_url": social.get("thumbnail_url", ""),
                "source_id": social.get("source_id", ""),
                "role": "Creator / producer",
                "publisher_client": social.get("account") or social.get("platform", ""),
                "priority_level": "low priority",
                "behind_the_story": "",
            }
        )

    deduped = {}
    for row in rows:
        if not (row.get("title") or row.get("url") or row.get("video_url")):
            continue
        key = source_key(row)
        current = deduped.get(key)
        if not current or platform_rank(row) < platform_rank(current):
            deduped[key] = row

    by_title = {}
    for row in deduped.values():
        if (row.get("content_type") or "video").lower() == "video":
            key = title_key(row) or source_key(row)
        else:
            key = source_key(row)
        current = by_title.get(key)
        if not current or platform_rank(row) < platform_rank(current):
            by_title[key] = row

    items = [to_item(row) for row in by_title.values()]

    def sort_key(item):
        priority = 0 if item["priority"] == "top priority" else 1
        year = int(item["year"]) if item["year"].isdigit() else 0
        return (priority, -year, item["type"], item["title"].lower())

    return sorted(items, key=sort_key)


def write_data(items, site_content):
    payload = {
        "site": site_content,
        "items": items,
        "filters": {
            "types": sorted({item["type"] for item in items}),
            "topics": sorted({topic for item in items for topic in item["topics"]}),
            "years": sorted(
                {item["year"] for item in items},
                key=lambda value: (value == "Undated", -(int(value) if value.isdigit() else 0)),
            ),
        },
    }
    DATA_JS.write_text("window.PORTFOLIO_DATA = " + json.dumps(payload, indent=2) + ";\n", encoding="utf-8")


def write_template():
    workbook = Workbook()
    workbook.remove(workbook.active)

    def add_sheet_to(target_workbook, name, headers, rows, widths=None):
        sheet = target_workbook.create_sheet(name)
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="111915")
        sheet.freeze_panes = "A2"
        for index, header in enumerate(headers, start=1):
            letter = sheet.cell(row=1, column=index).column_letter
            sheet.column_dimensions[letter].width = (widths or {}).get(header, max(14, min(60, len(header) + 10)))
        return sheet

    def add_sheet(name, headers, rows, widths=None):
        return add_sheet_to(workbook, name, headers, rows, widths)

    portfolio_headers = [
        "content_type",
        "platform",
        "account",
        "title",
        "date",
        "summary",
        "url",
        "thumbnail_url",
        "source_id",
        "role",
        "publisher_client",
        "priority_level",
        "behind_the_story",
        "topic",
        "notes",
    ]

    portfolio_rows = read_csv(PORTFOLIO_CSV)
    add_sheet("profile", ["field", "value"], PROFILE_ROWS, {"field": 24, "value": 90})
    add_sheet("experience", ["period", "title", "organization", "summary", "bullets"], EXPERIENCE_ROWS, {"period": 18, "title": 32, "organization": 32, "summary": 90, "bullets": 90})
    add_sheet("education", ["period", "title", "institution", "details"], EDUCATION_ROWS, {"period": 18, "title": 42, "institution": 48, "details": 60})
    add_sheet("skills", ["skill", "details"], SKILL_ROWS, {"skill": 28, "details": 80})
    add_sheet("partners", ["name", "url", "logo_path", "category"], PARTNER_ROWS, {"name": 30, "url": 52, "logo_path": 42, "category": 20})
    add_sheet("socials", ["platform", "url"], SOCIAL_ROWS, {"platform": 22, "url": 70})
    add_sheet("portfolio_content", portfolio_headers, portfolio_rows, {"title": 48, "summary": 70, "url": 58, "thumbnail_url": 58, "behind_the_story": 78, "notes": 38})
    add_sheet("social_videos", ["platform", "account", "title", "date", "summary", "video_url", "thumbnail_url", "source_id"], read_csv(SOCIAL_CSV), {"title": 55, "summary": 70, "video_url": 58, "thumbnail_url": 58})

    workbook.save(TEMPLATE_XLSX)
    if not CONTENT_XLSX.exists():
        workbook.save(CONTENT_XLSX)
    else:
        content_workbook = load_workbook(CONTENT_XLSX)
        existing = set(content_workbook.sheetnames)
        sheet_specs = [
            ("profile", ["field", "value"], PROFILE_ROWS, {"field": 24, "value": 90}),
            ("experience", ["period", "title", "organization", "summary", "bullets"], EXPERIENCE_ROWS, {"period": 18, "title": 32, "organization": 32, "summary": 90, "bullets": 90}),
            ("education", ["period", "title", "institution", "details"], EDUCATION_ROWS, {"period": 18, "title": 42, "institution": 48, "details": 60}),
            ("skills", ["skill", "details"], SKILL_ROWS, {"skill": 28, "details": 80}),
            ("partners", ["name", "url", "logo_path", "category"], PARTNER_ROWS, {"name": 30, "url": 52, "logo_path": 42, "category": 20}),
            ("socials", ["platform", "url"], SOCIAL_ROWS, {"platform": 22, "url": 70}),
            ("portfolio_content", portfolio_headers, portfolio_rows, {"title": 48, "summary": 70, "url": 58, "thumbnail_url": 58, "behind_the_story": 78, "notes": 38}),
            ("social_videos", ["platform", "account", "title", "date", "summary", "video_url", "thumbnail_url", "source_id"], read_csv(SOCIAL_CSV), {"title": 55, "summary": 70, "video_url": 58, "thumbnail_url": 58}),
        ]
        for name, headers, rows, widths in sheet_specs:
            if name not in existing:
                add_sheet_to(content_workbook, name, headers, rows, widths)
        content_workbook.save(CONTENT_XLSX)


if __name__ == "__main__":
    items = load_items()
    site_content = load_site_content()
    write_data(items, site_content)
    write_template()
    print(f"Wrote {len(items)} unique items to {DATA_JS.name}")
    print(f"Wrote template to {TEMPLATE_XLSX.name}")
